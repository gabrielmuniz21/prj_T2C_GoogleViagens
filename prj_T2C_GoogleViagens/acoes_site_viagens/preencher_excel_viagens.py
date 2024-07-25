import time
from selenium.webdriver.common.by import By
from botcity.web import WebBot, Browser
from openpyxl import load_workbook
import pandas as pd

class PreencherExcelViagens:
    def __init__(self, botWebbot:WebBot=None):
        self.var_botWebbot = botWebbot

    def salvarTodos(self, dados):
        caminho_excel_passagens = 'prj_T2C_GoogleViagens/Precos_viagem.xlsx'

        var_planilha_viagem = load_workbook(caminho_excel_passagens)
        var_planilha_editar = var_planilha_viagem.active
        
        # Seleciona a planilha 'Todos' ou cria uma nova se não existir
        if 'Todos' in var_planilha_viagem.sheetnames:
            var_planilha_editar = var_planilha_viagem['Todos']
        else:
            var_planilha_editar = var_planilha_viagem.create_sheet(title='Todos')

        # Encontrando a próxima linha vazia
        var_intIndexNewline = var_planilha_editar.max_row + 1

        # Adicionando os dados
        for dado in dados:
            var_planilha_editar.append([
                dado.get('País', ''),
                dado.get('Cidade', ''),
                dado.get('Preço', '')
            ])

        # Salvando e fechando o arquivo
        var_planilha_viagem.save(caminho_excel_passagens)
        var_planilha_viagem.close()
        
    def salvarBaratos(self):
        caminho_excel = 'prj_T2C_GoogleViagens/Precos_viagem.xlsx'
        var_planilha_precos = load_workbook(caminho_excel)
        
        df = pd.read_excel(caminho_excel, sheet_name='Todos')

        # Garantir que a coluna 'Preço' é interpretada corretamente (remover 'R$' e converter para numérico)
        df['Preço'] = df['Preço'].replace('[R$]', '', regex=True).astype(float)
        
        # Ordenar por preço e pegar as 10 linhas com os preços mais baixos
        df_sorted = df.sort_values(by='Preço').head(10)
        
        # Carregar o arquivo Excel para escrita
        with pd.ExcelWriter(caminho_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            var_planilha_precos = writer.book
            if 'Baratos' not in var_planilha_precos.sheetnames:
                var_wshtBaratos = var_planilha_precos.create_sheet(title='Baratos')
                var_wshtBaratos.append(['País', 'Cidade', 'Preço'])
            else:
                var_wshtBaratos = var_planilha_precos['Baratos']
                start_row = var_wshtBaratos.max_row + 1

            # Adiciona os dados na planilha 'Baratos'
            for row in df_sorted.itertuples(index=False, name=None):
                var_wshtBaratos.append(row)

            # Salvar o arquivo
            var_planilha_precos.save(caminho_excel)
    